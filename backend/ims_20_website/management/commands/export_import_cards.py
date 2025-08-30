# ////////////////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////////////////
# python manage.py export_import_cards export cards.xlsx
# python manage.py export_import_cards import cards.xlsx


# ////////////////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////////////////







# your_app/management/commands/export_import_cards.py
from django.core.management.base import BaseCommand
from openpyxl import Workbook, load_workbook
from django.db import transaction
from ims_20_website.models import CardItem, Feature, Institute


class Command(BaseCommand):
    help = "Export or Import CardItem & Feature records to/from Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "action",
            type=str,
            choices=["export", "import"],
            help="Action to perform: export or import",
        )
        parser.add_argument(
            "file_path",
            type=str,
            help="Path to Excel file (e.g., cards.xlsx)",
        )

    def handle(self, *args, **kwargs):
        action = kwargs["action"]
        file_path = kwargs["file_path"]

        if action == "export":
            self.export_cards(file_path)
        elif action == "import":
            self.import_cards(file_path)

    def export_cards(self, file_path):
        """Export all CardItems and Features into Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cards & Features"

        # Header row
        ws.append([
            "Institute Code",
            "Card Title",
            "Card Slug",
            "Card Icon",
            "Card Order",
            "Card Active",
            "Feature Text",
            "Feature Slug",
            "Feature Icon",
            "Feature Order",
            "Feature Active",
        ])

        for card in CardItem.objects.select_related("institute").prefetch_related("features"):
            for feature in card.features.all():
                ws.append([
                    card.institute.institute_code,
                    card.title,
                    card.slug,
                    card.icon,
                    card.order,
                    card.is_active,
                    feature.text,
                    feature.link,
                    feature.icon,
                    feature.order,
                    feature.is_active,
                ])

            if not card.features.exists():
                ws.append([
                    card.institute.institute_code,
                    card.title,
                    card.slug,
                    card.icon,
                    card.order,
                    card.is_active,
                    "", "", "", "", ""
                ])

        wb.save(file_path)
        self.stdout.write(self.style.SUCCESS(f"✅ Exported cards to {file_path}"))

    @transaction.atomic
    def import_cards(self, file_path):
        """Import/update CardItems and Features from Excel"""
        wb = load_workbook(filename=file_path)
        ws = wb.active

        count_cards = 0
        count_features = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                (
                    institute_code,
                    card_title,
                    card_slug,
                    card_icon,
                    card_order,
                    card_active,
                    feature_text,
                    feature_slug,
                    feature_icon,
                    feature_order,
                    feature_active,
                ) = row

                institute = Institute.objects.get(institute_code=institute_code)

                # Create/update CardItem
                card_obj, created = CardItem.objects.update_or_create(
                    title=card_title,
                    institute=institute,
                    defaults={
                        "slug": card_slug,
                        "icon": card_icon,
                        "order": card_order or 0,
                        "is_active": bool(card_active),
                    },
                )
                if created:
                    count_cards += 1

                # Create/update Feature (only if text is present)
                if feature_text:
                    feat_obj, f_created = Feature.objects.update_or_create(
                        card=card_obj,
                        text=feature_text,
                        defaults={
                            "link": feature_slug,
                            "icon": feature_icon,
                            "order": feature_order or 0,
                            "is_active": bool(feature_active),
                        },
                    )
                    if f_created:
                        count_features += 1

            except Exception as e:
                skipped += 1
                self.stdout.write(
                    self.style.WARNING(f"⚠️ Row {i} skipped due to error: {e}")
                )

        self.stdout.write(self.style.SUCCESS(f"✅ Imported/Updated {count_cards} cards"))
        self.stdout.write(self.style.SUCCESS(f"✅ Imported/Updated {count_features} features"))
        self.stdout.write(self.style.SUCCESS(f"⏭️ Skipped {skipped} invalid/error rows"))
