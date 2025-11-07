# Export all teachers to Excel
# python manage.py export_import_teachers export teachers.xlsx

# Import teachers from Excel (without images)
# python manage.py export_import_teachers import teachers.xlsx



import os
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from django.db import transaction
from ims_02_account.models import Teacher


class Command(BaseCommand):
    help = "Export or Import Teacher records to/from Excel (with images)"

    def add_arguments(self, parser):
        parser.add_argument(
            "action",
            type=str,
            choices=["export", "import"],
            help="Action to perform: export or import",
        )
        parser.add_argument(
            "file_path",
            type=str,
            help="Path to Excel file (e.g., teachers.xlsx)",
        )

    def handle(self, *args, **kwargs):
        action = kwargs["action"]
        file_path = kwargs["file_path"]

        if action == "export":
            self.export_teachers(file_path)
        elif action == "import":
            self.import_teachers(file_path)

    def export_teachers(self, file_path):
        """Export all teachers into Excel with images"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Teachers"

        # Header row
        ws.append([
            "ID", "Name", "Bangla Name", "Phone", "WhatsApp",
            "Designation", "Major Subject", "Blood Group",
            "Email", "Religion", "Section", "NID",
            "DOB", "Joining Date", "Indexing Date", "Index Number",
            "Qualification", "Address",
            "Picture", "Signature",
        ])

        row_num = 2
        for teacher in Teacher.objects.all().order_by("order"):
            ws.append([
                teacher.id,
                teacher.name,
                teacher.bangla_name or "",
                teacher.phone_number,
                "Yes" if teacher.is_whatsapp else "No",
                teacher.get_designation_display(),
                teacher.major_subject or "",
                teacher.blood_group or "",
                teacher.email or "",
                teacher.religion or "",
                teacher.section or "",
                teacher.nid or "",
                str(teacher.dob) if teacher.dob else "",
                str(teacher.joining_date) if teacher.joining_date else "",
                str(teacher.indexing_of_mpo) if teacher.indexing_of_mpo else "",
                teacher.index_number or "",
                teacher.qualification or "",
                teacher.address or "",
                "",  # placeholder for picture
                "",  # placeholder for signature
            ])

            # Insert images if available
            col_picture = "S"  # 19th column (Picture)
            col_signature = "T"  # 20th column (Signature)

            if teacher.picture and os.path.exists(teacher.picture.path):
                img = XLImage(teacher.picture.path)
                img.height = 80
                img.width = 60
                ws.add_image(img, f"{col_picture}{row_num}")

            if teacher.signature and os.path.exists(teacher.signature.path):
                img = XLImage(teacher.signature.path)
                img.height = 30
                img.width = 100
                ws.add_image(img, f"{col_signature}{row_num}")

            row_num += 1

        wb.save(file_path)
        self.stdout.write(self.style.SUCCESS(f"✅ Exported teachers to {file_path}"))

    @transaction.atomic
    def import_teachers(self, file_path):
        """
        Import/update teachers from Excel (⚠️ Note: does NOT import images,
        because openpyxl cannot easily map them back to file paths).
        """
        wb = load_workbook(filename=file_path)
        ws = wb.active

        count_teachers = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                (
                    teacher_id, name, bangla_name, phone, whatsapp,
                    designation, major_subject, blood_group,
                    email, religion, section, nid,
                    dob, joining_date, indexing_date, index_number,
                    qualification, address, _pic, _sign,
                ) = row

                Teacher.objects.update_or_create(
                    id=teacher_id,
                    defaults={
                        "name": name,
                        "bangla_name": bangla_name,
                        "phone_number": phone,
                        "is_whatsapp": str(whatsapp).lower() in ["yes", "1", "true"],
                        "designation": next(
                            (key for key, val in Teacher.DESIGNATION_CHOICES if val == designation),
                            "assistant_teacher"
                        ),
                        "major_subject": major_subject,
                        "blood_group": blood_group,
                        "email": email,
                        "religion": religion,
                        "section": section,
                        "nid": nid,
                        "dob": dob if dob else None,
                        "joining_date": joining_date if joining_date else None,
                        "indexing_of_mpo": indexing_date if indexing_date else None,
                        "index_number": index_number,
                        "qualification": qualification,
                        "address": address,
                    },
                )
                count_teachers += 1

            except Exception as e:
                skipped += 1
                self.stdout.write(
                    self.style.WARNING(f"⚠️ Row {i} skipped due to error: {e}")
                )

        self.stdout.write(self.style.SUCCESS(f"✅ Imported/Updated {count_teachers} teachers"))
        self.stdout.write(self.style.SUCCESS(f"⏭️ Skipped {skipped} invalid/error rows"))
