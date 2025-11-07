import os
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from django.db import transaction
from ims_02_account.models import Student
from ims_01_institute.models import Institute, Year, Class, Group, Section  # adjust paths



# class Command(BaseCommand):
#     help = "Export or import students to/from Excel"

#     def add_arguments(self, parser):
#         parser.add_argument("action", choices=["export", "import"], help="Action: export or import")
#         parser.add_argument("file_path", help="Path to the Excel file")

#     def handle(self, *args, **options):
#         action = options["action"]
#         file_path = options["file_path"]

#         if action == "export":
#             self.export_students(file_path)
#         elif action == "import":
#             self.stdout.write(self.style.WARNING("Import not implemented yet"))

#     def export_students(self, file_path):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Students"

#         # Header row
#         ws.append([
#             "ID", "Student ID", "Name", "Name (Bangla)", "Father's Name", "Mother's Name",
#             "Roll Number", "DOB", "Email", "Phone", "Guardian Phone", "Address",
#             "Institute", "Year", "Class", "Shift", "Group", "Section",
#         ])

#         students = Student.objects.select_related(
#             "institute", "year", "class_instance__class_name",
#             "group__group_name", "section__section_name"
#         )

#         for student in students:
#             ws.append([
#                 student.id,
#                 student.student_id,
#                 student.name,
#                 student.name_bangla,
#                 student.fathers_name,
#                 student.mothers_name,
#                 student.roll_number,
#                 student.dob.isoformat() if student.dob else "",
#                 student.email,
#                 student.phone_number,
#                 student.guardian_mobile_number,
#                 student.address,

#                 # Foreign Keys (make them human-readable)
#                 student.institute.name if student.institute else "",
#                 student.year.year if student.year else "",
#                 student.class_instance.class_name.name if student.class_instance and student.class_instance.class_name else "",
#                 student.class_instance.shift if student.class_instance else "",
#                 student.group.group_name.name if student.group and student.group.group_name else "",
#                 student.section.section_name.name if student.section and student.section.section_name else "",
#             ])

#         wb.save(file_path)
#         self.stdout.write(self.style.SUCCESS(f"Exported {students.count()} students to {file_path}"))


from django.core.management.base import BaseCommand
from ims_02_account.models import (
    Student, Institute, Year, Class, Group, Section
)
from openpyxl import Workbook, load_workbook
from django.db import transaction


class Command(BaseCommand):
    help = "Export or import students to/from Excel"

    def add_arguments(self, parser):
        parser.add_argument("action", choices=["export", "import"], help="Action: export or import")
        parser.add_argument("file_path", help="Path to the Excel file")

    def handle(self, *args, **options):
        action = options["action"]
        file_path = options["file_path"]

        if action == "export":
            self.export_students(file_path)
        elif action == "import":
            self.import_students(file_path)

    def export_students(self, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        # Header row
        ws.append([
            "ID", "Student ID", "Name", "Name (Bangla)", "Father's Name", "Mother's Name",
            "Roll Number", "DOB", "Email", "Phone", "Guardian Phone", "Address",
            "Institute", "Year", "Class", "Shift", "Group", "Section",
        ])

        students = Student.objects.select_related(
            "institute", "year", "class_instance__class_name",
            "group__group_name", "section__section_name"
        )

        for student in students:
            ws.append([
                student.id,
                student.student_id,
                student.name,
                student.name_bangla,
                student.fathers_name,
                student.mothers_name,
                student.roll_number,
                student.dob.isoformat() if student.dob else "",
                student.email,
                student.phone_number,
                student.guardian_mobile_number,
                student.address,

                student.institute.name if student.institute else "",
                student.year.year if student.year else "",
                student.class_instance.class_name.name if student.class_instance and student.class_instance.class_name else "",
                student.class_instance.shift if student.class_instance else "",
                student.group.group_name.name if student.group and student.group.group_name else "",
                student.section.section_name.name if student.section and student.section.section_name else "",
            ])

        wb.save(file_path)
        self.stdout.write(self.style.SUCCESS(f"Exported {students.count()} students to {file_path}"))

    @transaction.atomic
    def import_students(self, file_path):
        wb = load_workbook(file_path)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        data_rows = rows[1:]

        count = 0
        for row in data_rows:
            (
                _id, student_id, name, name_bangla, fathers_name, mothers_name,
                roll_number, dob, email, phone_number, guardian_mobile_number, address,
                institute_name, year_value, class_name, shift, group_name, section_name,
            ) = row

            if not name or not institute_name or not year_value:
                continue  # skip invalid rows

            try:
                institute = Institute.objects.get(name=institute_name)
            except Institute.DoesNotExist:
                self.stdout.write(self.style.WARNING(f"Skipping: Institute {institute_name} not found"))
                continue

            try:
                year = Year.objects.get(institute=institute, year=int(year_value))
            except Year.DoesNotExist:
                self.stdout.write(self.style.WARNING(f"Skipping: Year {year_value} not found for {institute_name}"))
                continue

            try:
                class_instance = Class.objects.get(
                    institute=institute, year=year,
                    class_name__name=class_name, shift=shift
                )
            except Class.DoesNotExist:
                self.stdout.write(self.style.WARNING(f"Skipping: Class {class_name} ({shift}) not found"))
                continue

            try:
                group = Group.objects.get(
                    institute=institute, year=year,
                    class_instance=class_instance,
                    group_name__name=group_name
                )
            except Group.DoesNotExist:
                group = None  # optional

            try:
                section = Section.objects.get(
                    institute=institute, year=year,
                    class_instance=class_instance,
                    group=group,
                    section_name__name=section_name
                )
            except Section.DoesNotExist:
                section = None  # optional

            student, created = Student.objects.update_or_create(
                student_id=student_id or f"{institute.id}-{year.year}-{roll_number}",
                defaults={
                    "name": name,
                    "name_bangla": name_bangla,
                    "fathers_name": fathers_name,
                    "mothers_name": mothers_name,
                    "roll_number": roll_number,
                    "dob": dob,
                    "email": email,
                    "phone_number": phone_number,
                    "guardian_mobile_number": guardian_mobile_number,
                    "address": address,
                    "institute": institute,
                    "year": year,
                    "class_instance": class_instance,
                    "group": group,
                    "section": section,
                }
            )
            count += 1

        self.stdout.write(self.style.SUCCESS(f"Imported {count} students from {file_path}"))
