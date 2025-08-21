from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ims_04_result.models import Institute, Year, Class, GroupName, Group

class Command(BaseCommand):
    help = 'Import Group records from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to Excel file (e.g., all_groups.xlsx)')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        wb = load_workbook(filename=file_path)
        ws = wb.active

        count = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                institute_code, year_val, class_name, shift, group_code, group_name, group_name_bn, order = row
                print("row: ", row)

                institute = Institute.objects.get(institute_code=institute_code)
                year = Year.objects.get(year=year_val)
                class_instance = Class.objects.get(institute=institute, year=year, class_name__name=class_name, shift=shift)
                # group_name_obj = GroupName.objects.get(code=group_code)
                
                group_name_obj, _ = GroupName.objects.update_or_create(
                    code=group_code,
                    defaults={
                        'name': group_name,
                        'name_bengali': group_name_bn or ""
                    }
                )

                Group.objects.update_or_create(
                    institute=institute,
                    year=year,
                    class_instance=class_instance,
                    group_name=group_name_obj,
                    defaults={'order': order}
                )
                count += 1

            except Exception as e:
                skipped += 1
                self.stdout.write(self.style.WARNING(f"⚠️ Row {i} skipped due to error: {e}"))

        self.stdout.write(self.style.SUCCESS(f"✅ Imported/Updated {count} group records"))
        self.stdout.write(self.style.SUCCESS(f"⏭️ Skipped {skipped} invalid or error rows"))
