from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ims_04_result.models import Section, SectionName, Institute, Year, Class, Group, ClassName

class Command(BaseCommand):
    help = 'Import Section records from Excel using string-based matching (Shift required)'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel file')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        wb = load_workbook(filename=file_path)
        ws = wb.active

        imported = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                (
                    institute_code,
                    year_val,
                    class_name_str,
                    shift,
                    group_name_str,
                    group_code,
                    section_code,
                    section_name,
                    section_bangla
                ) = row

                if not all([institute_code, year_val, class_name_str, shift, group_name_str,group_code, section_code, section_name]):
                    skipped += 1
                    self.stdout.write(self.style.WARNING(f"⚠️ Row {i} skipped: missing required fields"))
                    continue

                # Get related models
                institute = Institute.objects.get(institute_code=institute_code)
                year = Year.objects.get(year=year_val)
                class_name = ClassName.objects.get(name=class_name_str)
                class_obj = Class.objects.get(
                    institute=institute,
                    year=year,
                    class_name=class_name,
                    shift=shift
                )
                
                group = Group.objects.get(
                    class_instance=class_obj,
                    group_name__code=group_code
                )

                section_name_obj, _ = SectionName.objects.update_or_create(
                    code=section_code,
                    defaults={
                        'name': section_name,
                        'name_bengali': section_bangla or ""
                    }
                )

                Section.objects.update_or_create(
                    institute=institute,
                    year=year,
                    class_instance=class_obj,
                    group=group,
                    section_name=section_name_obj,
                )

                imported += 1

            except Exception as e:
                skipped += 1
                self.stdout.write(self.style.WARNING(f"⚠️ Row {i} skipped due to error: {e}"))

        self.stdout.write(self.style.SUCCESS(f"\n✅ Imported/Updated {imported} section records"))
        self.stdout.write(self.style.SUCCESS(f"⏭️ Skipped {skipped} invalid or error rows"))
