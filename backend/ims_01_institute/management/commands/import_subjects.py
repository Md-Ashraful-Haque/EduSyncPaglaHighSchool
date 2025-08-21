# //////////////////////////////////////////////////////////////////////////////////////////////////////////////
# Import Subject name using Excel file
# Command: python manage.py import_subjects subject_names_export_update.xlsx
# import_subjects is the python file name
# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////


from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ims_04_result.models import SubjectName  # adjust as needed

class Command(BaseCommand):
    help = 'Import SubjectName records from an Excel file (.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel file')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        wb = load_workbook(filename=file_path)
        ws = wb.active

        count = 0
        skipped = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Unpack row safely
            id_, name, name_bengali, short_name, class_name, code, serial = row

            # Skip rows where required fields are missing
            if not name or not code:
                skipped += 1
                self.stdout.write(self.style.WARNING(f"⚠️  Row {i} skipped: 'name' or 'code' is empty"))
                continue

            try:
                SubjectName.objects.update_or_create(
                    code=code,
                    defaults={
                        'name': name,
                        'name_bengali': name_bengali,
                        'short_name': short_name,
                        'class_name': class_name,
                        'serial': serial,
                    }
                )
                count += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f"⚠️  Row {i} error: {e}"))

        self.stdout.write(self.style.SUCCESS(f"\n✅ Imported/Updated {count} records"))
        self.stdout.write(self.style.SUCCESS(f"⏭️  Skipped {skipped} invalid or empty rows"))
