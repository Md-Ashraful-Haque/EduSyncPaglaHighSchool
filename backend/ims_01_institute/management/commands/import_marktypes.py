# //////////////////////////////////////////////////////////////////////////////////////////////////////////////
# Import Subject name using Excel file
# Command: python manage.py import_marktypes mark_types_export.xlsx
# import_subjects is the python file name
# ///////////////////////////////////////////////////////////////////////////////////////////////////////////////


from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ims_04_result.models import SubjectName, MarkTypeForSubject

class Command(BaseCommand):
    help = 'Import MarkTypeForSubject records from Excel using class_name and subject.code'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel (.xlsx) file')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        wb = load_workbook(filename=file_path)
        ws = wb.active

        count = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                class_name, subject_code, subject_name, subject_name_bengali, mark_type_display, max_marks, pass_marks, serial_number = row
                
                if not class_name or not subject_code or not mark_type_display:
                    skipped += 1
                    self.stdout.write(self.style.WARNING(f"⚠️  Row {i} skipped: Missing class, code, or mark type"))
                    continue

                # Reverse lookup mark_type value from display name
                mark_type_map = dict((v, k) for k, v in MarkTypeForSubject.MARK_TYPE_CHOICES)
                mark_type_key = mark_type_map.get(mark_type_display)

                if not mark_type_key:
                    skipped += 1
                    self.stdout.write(self.style.WARNING(f"⚠️  Row {i} skipped: Unknown mark type '{mark_type_display}'"))
                    continue

                subject = SubjectName.objects.filter(class_name=class_name, code=subject_code).first()

                if not subject:
                    skipped += 1
                    self.stdout.write(self.style.WARNING(f"⚠️  Row {i} skipped: Subject not found (Class: {class_name}, Code: {subject_code})"))
                    continue

                MarkTypeForSubject.objects.update_or_create(
                    subject=subject,
                    mark_type=mark_type_key,
                    defaults={
                        'max_marks': max_marks or 0,
                        'pass_marks': pass_marks or 0,
                        'serial_number': serial_number or 0
                    }
                )
                count += 1

            except Exception as e:
                skipped += 1
                self.stdout.write(self.style.WARNING(f"⚠️  Row {i} error: {e}"))

        self.stdout.write(self.style.SUCCESS(f"\n✅ Imported/Updated {count} records"))
        self.stdout.write(self.style.SUCCESS(f"⏭️  Skipped {skipped} rows due to issues"))
