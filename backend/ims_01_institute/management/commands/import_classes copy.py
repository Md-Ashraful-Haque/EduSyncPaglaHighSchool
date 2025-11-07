from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ims_04_result.models import Institute, Year, ClassName, Class

class Command(BaseCommand):
    help = 'Import Class records from Excel using class code to match ClassName'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to Excel file (e.g., all_classes.xlsx)')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        wb = load_workbook(filename=file_path)
        ws = wb.active

        count = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                institute_code, year_val, class_code, class_name, class_bn, shift = row
                print("row: ", row)

                institute = Institute.objects.get(institute_code=institute_code)
                year = Year.objects.get(year=year_val)
                # class_name_obj = ClassName.objects.get(code=class_code)

                class_name_obj,_ = ClassName.objects.update_or_create(
                    code=class_code,
                    defaults={
                        'name': class_name,
                        'name_bengali': class_bn or ""
                    }
                )
                print("class_name_obj: ",class_name_obj)
                
                SHIFT_CHOICES = [
                    ('morning', 'Morning'),
                    ('day', 'Day'),
                    ('evening', 'Evening'),
                ]

                SHIFT_CHOICES_DICT = {label.lower(): value for value, label in SHIFT_CHOICES}

                # inside import loop:
                raw_shift = str(shift).strip().lower()
                internal_shift = SHIFT_CHOICES_DICT.get(raw_shift)

                if not internal_shift:
                    raise ValueError(f"Invalid shift label: '{shift}'")

                Class.objects.update_or_create(
                    institute=institute,
                    year=year,
                    class_name=class_name_obj,
                    shift=internal_shift,
                )


                count += 1

            except Exception as e:
                skipped += 1
                self.stdout.write(self.style.WARNING(f"⚠️ Row {i} skipped due to error: {e}"))

        self.stdout.write(self.style.SUCCESS(f"✅ Imported/Updated {count} class records"))
        self.stdout.write(self.style.SUCCESS(f"⏭️ Skipped {skipped} invalid or error rows"))
